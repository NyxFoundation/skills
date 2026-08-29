#!/usr/bin/env python3
"""スケジュールと工数の唯一の出典。xlsx（週次ガント）と、提案書に貼る markdown 表を生成する。

**このファイル 1 本から xlsx と本文の表の両方を出すこと。** 本文の表を手で書くと必ずずれる。
実際に、本文の宣言値 425 人日と表の合計 447 人日が 3 版にわたり残った例がある。

下の PLANS / TASKS / MILESTONES を案件に合わせて書き換えて使う。
以下は「期間と費用の違う 3 案を、フェーズは同じで深さだけ変えて出す」場合の例である。

    uv run --with openpyxl python3 bin/make_schedule.py            # xlsx を書く
    uv run --with openpyxl python3 bin/make_schedule.py --md       # 提案書 12 章に貼る簡易表を出す

出力:
    out/schedule.xlsx   週次のガントチャート。プランごとに 1 シート
    標準出力（--md）    提案書に貼る簡易表（Markdown）

**このファイルがスケジュールと工数の唯一の出典である。** proposal.md 12 章の表は
`--md` の出力を貼ったものなので、直すときは必ずここを直して貼り直す。

3 プランの関係は「フェーズは同じ、深さが違う」である（2026-08-20 に入れ子構成から変更）。
どのプランでも 仕様 → 実装 → 検証 → 公開 の 4 フェーズを通し、成果物 D1〜D4 が揃う。
プランによって変えるのは対象の広さ（通貨ペア数、注文種別、Exposure の定義数）と
検証の深さであって、フェーズの有無ではない。プラン A でも公開まで到達する。
"""
import sys
from datetime import date, timedelta

START = date(2026, 10, 1)

PLANS = {
    "A": {"months": 3, "fee": "150 万円", "end": date(2026, 12, 30),
          "scope": "1 通貨ペア・指値のみ・Exposure 1 種類",
          "goal": "最小構成の Delegation Control Layer を作り、上限非超過を証明して公開する"},
    "B": {"months": 6, "fee": "300 万円", "end": date(2027, 3, 26),
          "scope": "複数ペア・成行を含む・Exposure 2 種類",
          "goal": "成行と部分約定を含む実運用に近い構成まで対象を広げ、実装との一致を検証する"},
    "C": {"months": 9, "fee": "500 万円", "end": date(2027, 6, 30),
          "scope": "B に加えて障害耐性・記録の再生・精度測定",
          "goal": "順序入替や欠落のある執行イベント下での挙動まで検証し、委譲枠の実用性を数値で示す"},
}

# (ID, タスク, フェーズ, 担当, 人日 A, 人日 B, 人日 C, 開始 A, 終了 A, 開始 B, 終了 B, 開始 C, 終了 C)
#   担当  N=自社主担当 / B=顧客主担当 / 共=共同
#   人日は 自社側の投入。顧客側は CLIENT_RATIO でタスク単位に積み上げる（下記）。
#   人日 0 のタスクはそのプランでは実施しない。フェーズ自体はどのプランにも存在する。
TASKS = [
    ("1.1", "キックオフ、体制と会議体の確定", "立ち上げ", "共", 1, 1, 1,
     "2026-10-01", "2026-10-02", "2026-10-01", "2026-10-02", "2026-10-01", "2026-10-02"),
    ("1.2", "共同研究契約・NDA 締結", "立ち上げ", "共", 3, 3, 3,
     "2026-10-01", "2026-10-16", "2026-10-01", "2026-10-16", "2026-10-01", "2026-10-16"),
    ("1.3", "成果物の公開範囲および IP の合意", "立ち上げ", "共", 2, 2, 2,
     "2026-10-07", "2026-10-16", "2026-10-07", "2026-10-16", "2026-10-07", "2026-10-16"),
    ("1.4", "環境アクセス付与（リポジトリ、モック、API）", "立ち上げ", "B", 0, 0, 0,
     "2026-10-07", "2026-10-16", "2026-10-07", "2026-10-16", "2026-10-07", "2026-10-16"),
    ("1.5", "14 章の確認事項への回答とすり合わせ", "立ち上げ", "共", 3, 3, 3,
     "2026-10-07", "2026-10-23", "2026-10-07", "2026-10-23", "2026-10-07", "2026-10-23"),

    ("2.1", "Exposure 定義の選択と決定", "仕様策定", "共", 4, 6, 6,
     "2026-10-14", "2026-10-30", "2026-10-14", "2026-10-30", "2026-10-14", "2026-10-30"),
    ("2.2", "執行イベントモデルの形式化", "仕様策定", "N", 3, 6, 8,
     "2026-10-21", "2026-10-30", "2026-10-21", "2026-11-06", "2026-10-21", "2026-11-06"),
    ("2.3", "決定を要する論点の検討と決定", "仕様策定", "共", 4, 8, 10,
     "2026-10-28", "2026-11-06", "2026-10-28", "2026-11-13", "2026-10-28", "2026-11-13"),
    ("2.4", "監査ログの記録項目の設計と合意", "仕様策定", "共", 2, 4, 5,
     "2026-11-04", "2026-11-13", "2026-11-04", "2026-11-20", "2026-11-04", "2026-11-20"),
    ("2.5", "委譲仕様書 v1.0 の統合とレビュー（D1）", "仕様策定", "N", 3, 5, 5,
     "2026-11-11", "2026-11-20", "2026-11-18", "2026-11-27", "2026-11-18", "2026-11-27"),

    ("3.1", "Cedar スキーマとポリシーカタログ", "実装", "N", 4, 8, 8,
     "2026-10-28", "2026-11-13", "2026-11-04", "2026-12-04", "2026-11-04", "2026-12-04"),
    ("3.2", "予算状態機械の実装（TypeScript）", "実装", "N", 7, 12, 12,
     "2026-11-04", "2026-12-25", "2026-11-18", "2026-12-25", "2026-11-18", "2026-11-27"),
    ("3.3", "Cedar 統合（cedar-wasm）", "実装", "N", 3, 4, 4,
     "2026-11-18", "2026-11-30", "2026-12-09", "2026-12-28", "2026-12-09", "2026-12-28"),
    ("3.4", "執行イベントの購読と冪等な状態更新", "実装", "N", 4, 8, 10,
     "2026-11-18", "2026-12-04", "2026-12-16", "2027-01-15", "2026-12-16", "2027-01-15"),
    ("3.5", "監査ログの実装と再生機構", "実装", "N", 3, 5, 7,
     "2026-11-25", "2026-12-11", "2026-12-30", "2027-01-22", "2026-12-30", "2027-01-22"),
    ("3.6", "テストコードと利用ドキュメントの整備（D2）", "実装", "N", 3, 6, 8,
     "2026-12-02", "2026-12-18", "2027-01-13", "2027-02-12", "2027-01-13", "2027-02-12"),
    ("3.7", "実験環境の準備（モック拡張、顧客側）", "実装", "B", 0, 0, 0,
     "2026-11-04", "2026-11-27", "2026-11-18", "2027-01-01", "2026-11-18", "2027-01-01"),
    ("3.8", "MCP と取引 API の間への接続と結合試験", "実装", "共", 3, 4, 5,
     "2026-12-09", "2026-12-18", "2027-02-03", "2027-02-19", "2027-02-03", "2027-02-19"),
    ("3.9", "動作確認画面の実装", "実装", "N", 0, 5, 10,
     "2026-12-01", "2026-12-01", "2027-02-10", "2027-03-05", "2027-03-17", "2027-04-23"),

    ("4.1", "Lean による状態機械の形式仕様", "検証", "N", 4, 8, 8,
     "2026-11-11", "2026-11-27", "2026-12-02", "2027-01-01", "2026-12-02", "2027-01-01"),
    ("4.2", "不変量の証明（定理 A、D3）", "検証", "N", 7, 14, 16,
     "2026-11-25", "2026-12-18", "2026-12-30", "2027-02-19", "2026-12-30", "2027-02-19"),
    ("4.3", "Cedar symcc によるポリシー解析", "検証", "N", 2, 3, 3,
     "2026-12-02", "2026-12-11", "2027-01-06", "2027-01-22", "2027-01-06", "2027-01-22"),
    ("4.4", "差分テスト基盤の構築と実行", "検証", "N", 2, 8, 10,
     "2026-12-09", "2026-12-25", "2027-02-03", "2027-03-05", "2027-02-03", "2027-03-05"),
    ("4.5", "記録の忠実性の証明（定理 B）", "検証", "N", 0, 4, 6,
     "2026-12-01", "2026-12-01", "2027-02-24", "2027-03-12", "2027-02-24", "2027-03-12"),
    ("4.6", "障害注入と順序入替に対する挙動の検証", "検証", "共", 0, 3, 8,
     "2026-12-01", "2026-12-01", "2027-03-03", "2027-03-19", "2027-03-17", "2027-04-16"),
    ("4.7", "精度指標（過剰拒否率・τ・実効利用率）の測定", "検証", "N", 0, 0, 8,
     "2026-12-01", "2026-12-01", "2026-12-01", "2026-12-01", "2027-04-14", "2027-05-07"),
    ("4.8", "証明の前提条件の文書化", "検証", "N", 2, 3, 3,
     "2026-12-16", "2026-12-25", "2027-03-10", "2027-03-19", "2027-05-05", "2027-05-14"),

    ("5.1", "公開範囲と公開先の決定", "公開", "共", 1, 2, 2,
     "2026-12-09", "2026-12-18", "2027-02-24", "2027-03-05", "2027-04-21", "2027-04-30"),
    ("5.2", "研究成果の執筆（D4）", "公開", "共", 5, 10, 14,
     "2026-12-09", "2026-12-25", "2027-02-24", "2027-03-19", "2027-04-28", "2027-06-11"),
    ("5.3", "OSS 整備とリポジトリ公開", "公開", "N", 2, 4, 5,
     "2026-12-16", "2026-12-30", "2027-03-10", "2027-03-26", "2027-05-19", "2027-06-18"),
    ("5.4", "成果報告会", "公開", "共", 1, 1, 2,
     "2026-12-23", "2026-12-30", "2027-03-17", "2027-03-26", "2027-06-16", "2027-06-30"),
]

MILESTONES = {
    "A": [("M1", "仕様確定", "2026-11-20", "Exposure 定義と論点が決定し、D1 を受領"),
          ("M2", "実装・検証完了", "2026-12-25", "D2 と D3 が揃い、上限非超過が証明されている"),
          ("M3", "公開完了", "2026-12-30", "D4 を公開し、成果報告会を実施")],
    "B": [("M1", "仕様確定", "2026-11-27", "Exposure 定義と論点が決定し、D1 を受領"),
          ("M2", "実装完了", "2027-03-05", "D2 が揃い、実験環境に対して動作"),
          ("M3", "検証完了", "2027-03-19", "D3 が揃い、差分テストが一致"),
          ("M4", "公開完了", "2027-03-26", "D4 を公開し、成果報告会を実施")],
    "C": [("M1", "仕様確定", "2026-11-27", "Exposure 定義と論点が決定し、D1 を受領"),
          ("M2", "実装完了", "2027-04-23", "D2 が揃い、実験環境に対して動作"),
          ("M3", "検証完了", "2027-05-14", "D3 が揃い、障害耐性と精度指標の測定が完了"),
          ("M4", "公開完了", "2027-06-30", "D4 を公開し、成果報告会を実施")],
}

PLAN_IDX = {"A": 0, "B": 1, "C": 2}
PHASES = ["立ち上げ", "仕様策定", "実装", "検証", "公開"]
PHASE_COLOR = {
    "立ち上げ": "94A3B8", "仕様策定": "4F46E5", "実装": "0891B2",
    "検証": "7C3AED", "公開": "BE185D",
}
# 顧客側の投入見込みは、担当区分ごとの比率でタスクから積み上げる。別枠で手入力しない。
#   共 … 共同作業。自社が主導し顧客が意思決定と実務知識を出す想定で 0.5
#   N  … 自社主担当。レビュー分として 0.10
#   B  … 顧客主担当。自社側の人日が 0 なので CLIENT_TASK_DAYS で別建てにする
CLIENT_RATIO = {"共": 0.5, "B": 0.0, "N": 0.10}

# 顧客主担当タスクの 顧客側工数（自社側人日が 0 なので個別に持つ）
CLIENT_TASK_DAYS = {
    "1.4": {"A": 2, "B": 2, "C": 2},      # 環境アクセス付与
    "3.7": {"A": 4, "B": 10, "C": 14},    # 実験環境の準備
}

# 予備工数の比率。作業工数に対して上乗せし、フェーズには配分しない。
BUFFER_RATIO = 0.15


def d(s):
    return date.fromisoformat(s)


def effort(task, plan):
    return task[4 + PLAN_IDX[plan]]


def span(task, plan):
    i = 7 + PLAN_IDX[plan] * 2
    return d(task[i]), d(task[i + 1])


def tasks_for(plan):
    return [t for t in TASKS if effort(t, plan) > 0 or t[3] == "B"]


def own_days(plan):
    """作業工数（予備工数を含まない）。"""
    return sum(effort(t, plan) for t in TASKS)


def own_buffer(plan):
    """予備工数。小数点以下は切り上げる。"""
    import math
    return math.ceil(own_days(plan) * BUFFER_RATIO)


def own_total(plan):
    return own_days(plan) + own_buffer(plan)


def client_effort(task, plan):
    """タスク 1 件あたりの 顧客側工数。小数第 1 位で丸める。"""
    if task[0] in CLIENT_TASK_DAYS:
        return float(CLIENT_TASK_DAYS[task[0]][plan])
    e = effort(task, plan)
    if e <= 0:
        return 0.0
    return round(e * CLIENT_RATIO[task[3]], 1)


def client_days(plan):
    """顧客側の投入見込み。タスク単位の積み上げを小数点以下切り上げ。"""
    import math
    return math.ceil(sum(client_effort(t, plan) for t in TASKS))


def weeks(plan):
    end = PLANS[plan]["end"]
    w, cur = [], START - timedelta(days=START.weekday())
    while cur <= end:
        w.append(cur)
        cur += timedelta(days=7)
    return w


# ---------------------------------------------------------------- Markdown
def emit_md():
    print("### プラン比較\n")
    print("| | プラン A | プラン B | プラン C |")
    print("|---|---|---|---|")
    print("| 期間 | " + " | ".join(f"{PLANS[p]['months']} か月" for p in "ABC") + " |")
    print("| 費用 | " + " | ".join(PLANS[p]["fee"] for p in "ABC") + " |")
    print("| 終期 | " + " | ".join(PLANS[p]["end"].isoformat() for p in "ABC") + " |")
    print("| 対象の広さ | " + " | ".join(PLANS[p]["scope"] for p in "ABC") + " |")
    print("| 到達点 | " + " | ".join(PLANS[p]["goal"] for p in "ABC") + " |")
    print("| 成果物 | " + " | ".join("D1〜D4（すべて）" for p in "ABC") + " |")
    print("| 自社 作業工数 | " + " | ".join(f"{own_days(p)} 人日" for p in "ABC") + " |")
    print("| 自社 予備工数（15%） | " + " | ".join(f"{own_buffer(p)} 人日" for p in "ABC") + " |")
    print("| 自社 投入見込み 合計 | " + " | ".join(f"**{own_total(p)} 人日**" for p in "ABC") + " |")
    print("| 顧客 投入見込み | " + " | ".join(f"{client_days(p)} 人日" for p in "ABC") + " |")

    print("\n### フェーズごとの深さ\n")
    print("| フェーズ | プラン A | プラン B | プラン C |")
    print("|---|---|---|---|")
    for ph in PHASES:
        cells = []
        for p in "ABC":
            ts = [t for t in TASKS if t[2] == ph and effort(t, p) > 0]
            if not ts:
                cells.append("—")
                continue
            s = min(span(t, p)[0] for t in ts)
            e = max(span(t, p)[1] for t in ts)
            cells.append(f"{s.strftime('%m/%d')} 〜 {e.strftime('%m/%d')}（{sum(effort(t, p) for t in ts)} 人日）")
        print(f"| {ph} | " + " | ".join(cells) + " |")

    print("\n### マイルストーン\n")
    for p in "ABC":
        print(f"**プラン {p}**\n")
        print("| ID | 時期 | 内容 | 通過判定 |")
        print("|---|---|---|---|")
        for mid, name, when, gate in MILESTONES[p]:
            print(f"| {mid} | {when} | {name} | {gate} |")
        print()


# ---------------------------------------------------------------- xlsx
def emit_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="D6DBE3")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    for plan in "ABC":
        p = PLANS[plan]
        ws = wb.create_sheet(f"プラン{plan}（{p['months']}か月）")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "G4"

        head = ["#", "タスク", "フェーズ", "担当", "人日"]
        wk = weeks(plan)

        ws.cell(1, 1, f"プラン{plan}　{p['months']} か月 / {p['fee']}").font = Font(bold=True, size=14)
        ws.cell(2, 1, f"到達点: {p['goal']}").font = Font(size=11, color="5B6472")
        ws.cell(2, 4, f"対象: {p['scope']}").font = Font(size=11, color="5B6472")
        ws.cell(1, 5, f"自社 {own_days(plan)}+予備{own_buffer(plan)}={own_total(plan)} 人日 / 顧客 {client_days(plan)} 人日").font = Font(
            bold=True, size=11, color="5B6472")

        for c, h in enumerate(head, start=1):
            cell = ws.cell(3, c, h)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="374151")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = grid
        for i, w in enumerate(wk):
            cell = ws.cell(3, len(head) + 1 + i, w.strftime("%m/%d"))
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="374151")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = grid

        r = 4
        for t in tasks_for(plan):
            tid, name, phase, owner = t[0], t[1], t[2], t[3]
            nd = effort(t, plan)
            s, e = span(t, plan)
            ws.cell(r, 1, tid).alignment = Alignment(horizontal="center")
            ws.cell(r, 2, name)
            ws.cell(r, 3, phase).alignment = Alignment(horizontal="center")
            ws.cell(r, 4, owner).alignment = Alignment(horizontal="center")
            ws.cell(r, 5, nd or "").alignment = Alignment(horizontal="center")
            for c in range(1, len(head) + 1):
                ws.cell(r, c).border = grid
                ws.cell(r, c).font = Font(size=10)
            fill = PatternFill("solid", fgColor=PHASE_COLOR[phase])
            for i, w in enumerate(wk):
                cell = ws.cell(r, len(head) + 1 + i)
                cell.border = grid
                if s < w + timedelta(days=7) and e >= w:
                    cell.fill = fill
            r += 1

        r += 1
        ws.cell(r, 2, "マイルストーン").font = Font(bold=True, size=11)
        r += 1
        for mid, mname, when, gate in MILESTONES[plan]:
            ws.cell(r, 1, mid).alignment = Alignment(horizontal="center")
            ws.cell(r, 2, f"{mname}　{when}")
            ws.cell(r, 3, gate)
            for c in (1, 2, 3):
                ws.cell(r, c).font = Font(size=10, bold=(c == 2))
            for i, w in enumerate(wk):
                if d(when) < w + timedelta(days=7) and d(when) >= w:
                    cell = ws.cell(r, len(head) + 1 + i, "◆")
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=11, bold=True, color="B91C1C")
            r += 1

        for c, wdt in enumerate([6, 44, 10, 6, 8], start=1):
            ws.column_dimensions[get_column_letter(c)].width = wdt
        for i in range(len(wk)):
            ws.column_dimensions[get_column_letter(len(head) + 1 + i)].width = 4.2

    wb.save(path)
    print("saved:", path, f"({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    if "--md" in sys.argv:
        emit_md()
    else:
        import os
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        emit_xlsx(os.path.join(base, "out", "schedule.xlsx"))
