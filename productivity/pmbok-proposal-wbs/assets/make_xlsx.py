# -*- coding: utf-8 -*-
"""WBS / Gantt / estimate workbook skeleton.
    uv run --quiet --with openpyxl python make_xlsx.py
Replace the DATA section. Everything above it is the reusable machinery.
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

JP = "Yu Gothic"
OUT = "wbs_gantt.xlsx"

C_TITLE = C_HDR = "1F3864"
C_HDR_TXT = "FFFFFF"
C_SUB = "D9E2F3"
C_INPUT = "FFF2CC"          # 入力欄（黄）— 入力欄はここだけ、と凡例で示す
C_NOTE = "F2F2F2"
C_GATE = "C00000"
PHASE_COLOR = {"P0": "8FAADC", "P1": "A9D18E", "P2": "FFD966", "P3": "F4B183",
               "P4": "C5A5CF", "P5": "9DC3E6", "P6": "BFBFBF", "PX": "E2EFDA"}
PHASE_NAME = {"P0": "P0 準備", "P1": "P1 …", "P2": "P2 …", "P3": "P3 …",
              "P4": "P4 …", "P5": "P5 …", "P6": "P6 報告", "PX": "PX 横断"}

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(c, *, bold=False, size=9, color="000000", fill=None,
               align="left", wrap=False, border=True, fmt=None):
    c.font = Font(name=JP, size=size, bold=bold, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    if fmt:
        c.number_format = fmt
    return c


def put(ws, row, col, value, **kw):
    return style_cell(ws.cell(row=row, column=col, value=value), **kw)


# ── 営業日カレンダー ────────────────────────────────────────
HOLIDAYS = {dt.date(2026, 10, 12)}      # 着手日を変えたら祝日を再計算すること
START = dt.date(2026, 10, 5)
NDAYS = 19


def build_days(start, n):
    days, d = [], start
    while len(days) < n:
        if d.weekday() < 5 and d not in HOLIDAYS:
            days.append(d)
        d += dt.timedelta(days=1)
    return days


DAYS = build_days(START, NDAYS)
WEEK_OF = [(d - DAYS[0]).days // 7 + 1 for d in DAYS]

# ═══════════════════════════════════════════ DATA — replace below
# (ID, フェーズ, 種別, 作業内容, 主担当, 協力, 人日, 開始D, 終了D, 成果物, 完了条件)
ROWS = [
    ("P0.1", "P0", "タスク", "キックオフ（目的・成功基準・体制の合意）", "PMO", "全員",
     1.0, 1, 1, "議事録", "成功判定基準に全員が同意し議事録に記名"),
    ("P0.2", "P0", "タスク", "対象資料の受領", "顧客", "受注側",
     0.25, 1, 1, "受領物一覧", "アクセス払い出し完了"),
    ("G0", "P0", "ゲート", "【ゲート】スコープ凍結", "顧客", "",
     0.0, 2, 2, "承認記録", "スコープ承認／対象凍結／環境確認"),
    ("P1.1", "P1", "タスク", "……", "受注側", "",
     3.0, 2, 5, "……", "……"),
    ("G1", "P1", "ゲート", "【ゲート】……", "顧客", "",
     0.0, 5, 5, "承認記録", "……"),
    ("PX.1", "PX", "横断", "週次定例", "PMO", "", 2.0, 1, NDAYS, "週次報告", "進捗・課題・翌週予定を共有"),
    # 予備工数は 1 行として計上し、フェーズには配分しない
    ("PX.9", "PX", "予備工数", "予備工数（作業工数の約 15%。PM が一括管理）", "PMO", "",
     6.5, 1, NDAYS, "—", "消費した場合は理由とともに週次で報告"),
]

# ロール別工数は WBS の受注側積み上げと一致させること（capacity 検算も必須）
ROLES = [
    ("受注側", "リード",        0.79, "4 週（19 日）", 15.0),
    ("受注側", "エンジニア A",  0.89, "3 週（14 日）", 12.5),
    ("受注側", "エンジニア B",  0.89, "3 週（14 日）", 12.5),
    ("PMO",    "PM",            0.53, "4 週（19 日）", 10.0),
]
CLIENT_EFFORT = "開発 X.0 人日／セキュリティ X.0 人日（費用対象外）"
# ═══════════════════════════════════════════ END DATA

wb = Workbook()

# ───────────────────────────────── Sheet: WBS・ガント
gw = wb.active
gw.title = "WBS・ガント"
gw.sheet_view.showGridLines = False
HDRS = ["ID", "フェーズ", "種別", "作業内容", "主担当", "協力", "人日", "開始", "終了",
        "主な成果物", "完了条件（DoD）"]
for i, w in enumerate([7, 6, 8, 46, 11, 12, 6, 5, 5, 24, 34], start=1):
    gw.column_dimensions[get_column_letter(i)].width = w
GC0 = len(HDRS) + 1                      # first Gantt column
for i in range(NDAYS):
    gw.column_dimensions[get_column_letter(GC0 + i)].width = 3.6

put(gw, 1, 1, "WBS / ガントチャート", bold=True, size=14, color=C_TITLE, border=False)
put(gw, 2, 1, "凡例：帯＝作業期間（フェーズ別に着色）／赤＝承認ゲート。開始・終了は営業日通番。",
    size=9, color="595959", border=False)

HR = 5                                   # header row
wk = {}
for i, w in enumerate(WEEK_OF):
    wk.setdefault(w, []).append(i)
for w, idxs in wk.items():
    c0, c1 = GC0 + idxs[0], GC0 + idxs[-1]
    gw.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c1)
    put(gw, 3, c0, f"第{w}週", bold=True, size=9, fill=C_SUB, align="center")
    for c in range(c0, c1 + 1):
        gw.cell(row=3, column=c).border = BOX
for i, d in enumerate(DAYS):
    put(gw, 4, GC0 + i, d.strftime("%m/%d"), size=7, align="center", fill=C_NOTE)
    gw.cell(row=4, column=GC0 + i).alignment = Alignment(
        horizontal="center", vertical="center", textRotation=90)
gw.row_dimensions[4].height = 34
for i, h in enumerate(HDRS, start=1):
    put(gw, HR, i, h, bold=True, size=9, fill=C_HDR, color=C_HDR_TXT, align="center", wrap=True)
for i in range(NDAYS):
    c = put(gw, HR, GC0 + i, i + 1, bold=True, size=8, fill=C_HDR, color=C_HDR_TXT, align="center")
    c.number_format = '"D"0'             # numeric value so the CF formula can compare it
gw.row_dimensions[HR].height = 26

row = HR + 1
cur = None
for rid, ph, kind, task, own, sup, md, s, e, deliv, dod in ROWS:
    if ph != cur:                        # phase band (do NOT merge across the Gantt columns)
        cur = ph
        gw.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HDRS))
        put(gw, row, 1, PHASE_NAME[ph], bold=True, size=10, fill=PHASE_COLOR[ph])
        for c in range(1, len(HDRS) + 1):
            gw.cell(row=row, column=c).border = BOX
        for i in range(NDAYS):
            style_cell(gw.cell(row=row, column=GC0 + i), fill=PHASE_COLOR[ph])
        row += 1
    gate = kind == "ゲート"
    put(gw, row, 1, rid, size=9, align="center", bold=gate)
    put(gw, row, 2, ph, size=9, align="center")
    put(gw, row, 3, kind, size=9, align="center")
    put(gw, row, 4, task, size=9, wrap=True, bold=gate, color=C_GATE if gate else "000000")
    put(gw, row, 5, own, size=9, align="center")
    put(gw, row, 6, sup, size=9, align="center")
    put(gw, row, 7, md, size=9, align="center", fmt="0.00")
    put(gw, row, 8, s, size=9, align="center", fmt='"D"0')
    put(gw, row, 9, e, size=9, align="center", fmt='"D"0')
    put(gw, row, 10, deliv, size=9, wrap=True)
    put(gw, row, 11, dod, size=9, wrap=True)
    for i in range(NDAYS):
        style_cell(gw.cell(row=row, column=GC0 + i), size=8)
    gw.row_dimensions[row].height = 26
    row += 1
LAST = row - 1

put(gw, row, 1, "合計", bold=True, size=10, fill=C_SUB, align="center")
gw.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for c in range(1, 7):
    gw.cell(row=row, column=c).border = BOX
put(gw, row, 7, f"=SUM(G{HR+1}:G{LAST})", bold=True, size=10, align="center", fill=C_SUB, fmt="0.00")
for i in range(8, len(HDRS) + NDAYS + 1):
    style_cell(gw.cell(row=row, column=i), fill=C_SUB)
put(gw, row + 2, 1,
    "※ 上表はタスク単位の積み上げで、お客様側の稼働も含みます。"
    "ご請求の基礎となるロール別の内訳は「工数・見積」シートをご参照ください。",
    size=9, color="595959", border=False)

# Gantt bars via conditional formatting so they follow edits to 開始/終了
gr = f"{get_column_letter(GC0)}{HR+1}:{get_column_letter(GC0+NDAYS-1)}{LAST}"
L = get_column_letter(GC0)
gw.conditional_formatting.add(gr, FormulaRule(
    formula=[f'AND($C{HR+1}="ゲート",{L}${HR}>=$H{HR+1},{L}${HR}<=$I{HR+1})'],
    fill=PatternFill("solid", fgColor=C_GATE), stopIfTrue=True))
for ph, color in PHASE_COLOR.items():
    gw.conditional_formatting.add(gr, FormulaRule(
        formula=[f'AND($B{HR+1}="{ph}",{L}${HR}>=$H{HR+1},{L}${HR}<=$I{HR+1})'],
        fill=PatternFill("solid", fgColor=color), stopIfTrue=True))

gw.freeze_panes = f"{get_column_letter(GC0)}{HR+1}"      # string, not a cell object
gw.auto_filter.ref = f"A{HR}:{get_column_letter(len(HDRS))}{LAST}"
gw.print_title_rows = f"{HR}:{HR}"
gw.page_setup.orientation = "landscape"
gw.page_setup.fitToWidth = 1
gw.sheet_properties.pageSetUpPr.fitToPage = True

# ───────────────────────────────── Sheet: 工数・見積
es = wb.create_sheet("工数・見積")
es.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHI", [3, 30, 14, 10, 10, 10, 16, 18, 18]):
    es.column_dimensions[col].width = w

put(es, 2, 2, "お見積り（工数・費用）", bold=True, size=16, color=C_TITLE, border=False)
put(es, 3, 2, "黄色のセルが入力欄です（人月単価・予備工数比率・人月あたり稼働日数）。",
    size=10, color="595959", border=False)

r = 5
put(es, r, 2, "人月あたり稼働日数", size=10, bold=True)
MD_PER_MM = f"$C${r}"
put(es, r, 3, 20, size=10, align="center", fill=C_INPUT, color="0000FF")
put(es, r, 4, "人日／人月", size=10)
r += 2

hdr = ["担当", "ロール", "稼働率", "期間", "人日", "人月", "人月単価（円）", "金額（円）"]
for i, h in enumerate(hdr):
    put(es, r, 2 + i, h, bold=True, size=9, fill=C_HDR, color=C_HDR_TXT, align="center", wrap=True)
r += 1
first = r
for who, role, rate, per, md in ROLES:
    put(es, r, 2, who, size=10, align="center")
    put(es, r, 3, role, size=10)
    put(es, r, 4, rate, size=10, align="center", fmt="0%")
    put(es, r, 5, per, size=10, align="center")
    put(es, r, 6, md, size=10, align="center", fmt="0.0")
    put(es, r, 7, f"=F{r}/{MD_PER_MM}", size=10, align="center", fmt="0.00")
    put(es, r, 8, 0, size=10, align="right", fill=C_INPUT, color="0000FF", fmt="#,##0")
    put(es, r, 9, f"=G{r}*H{r}", size=10, align="right", fmt="#,##0")
    r += 1
last = r - 1

put(es, r, 2, "作業工数 小計", bold=True, size=10, fill=C_SUB, align="center")
put(es, r, 6, f"=SUM(F{first}:F{last})", bold=True, size=10, align="center", fill=C_SUB, fmt="0.0")
put(es, r, 7, f"=SUM(G{first}:G{last})", bold=True, size=10, align="center", fill=C_SUB, fmt="0.00")
put(es, r, 9, f"=SUM(I{first}:I{last})", bold=True, size=10, align="right", fill=C_SUB, fmt="#,##0")
BASE = r
r += 1

put(es, r, 2, "予備工数", bold=True, size=10, align="center")
put(es, r, 3, "前提条件の揺れに対する予備。フェーズに配分せず PM が一括管理", size=10)
put(es, r, 5, 0.15, size=10, align="center", fill=C_INPUT, color="0000FF", fmt="0%")
RATE = f"$E${r}"
# round to the nearest half day so the workbook agrees with the document
put(es, r, 6, f"=ROUND(F{BASE}*{RATE}*2,0)/2", size=10, align="center", fmt="0.0")
put(es, r, 7, f"=F{r}/{MD_PER_MM}", size=10, align="center", fmt="0.00")
put(es, r, 9, f"=G{r}*AVERAGE(H{first}:H{last})", size=10, align="right", fmt="#,##0")
BUF = r
r += 1

put(es, r, 2, "合計（予備工数込み）", bold=True, size=10, fill=C_SUB, align="center")
put(es, r, 6, f"=F{BASE}+F{BUF}", bold=True, size=10, align="center", fill=C_SUB, fmt="0.0")
put(es, r, 7, f"=G{BASE}+G{BUF}", bold=True, size=10, align="center", fill=C_SUB, fmt="0.00")
put(es, r, 9, f"=I{BASE}+I{BUF}", bold=True, size=11, align="right", fill=C_SUB, fmt="#,##0")
r += 2
put(es, r, 2, f"※ 上表は「WBS・ガント」シートの受注側タスク積み上げと一致します。"
              f"このほかにお客様側のご稼働として {CLIENT_EFFORT}。",
    size=9, color="595959", border=False)
r += 1
put(es, r, 2, "※ 予備工数の金額は、ロール別人月単価の平均で算出しています。",
    size=9, color="595959", border=False)

wb.save(OUT)
print("saved", OUT)
